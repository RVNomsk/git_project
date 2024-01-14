import os
import sqlite3
import openpyxl


def export_to_sqlite():
    """Экспорт данных из xlsx в sqlite"""

    # Текущая папка
    prj_dir = os.path.abspath(os.path.curdir)
    # a = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

    # SQL name БД
    base_name = 'coffee.sqlite'

    # метод sqlite3.connect автоматически создаст базу, если ее нет
    connect = sqlite3.connect(prj_dir + '/' + base_name)
    # курсор - это специальный объект, который делает запросы и
    # получает результаты запросов
    cursor = connect.cursor()

    # создание таблицы если ее не существует
    cursor.execute("""CREATE TABLE IF NOT EXISTS Coffee
                (id INTEGER PRIMARY KEY,
                variety_name TEXT,
                degree_of_roasting TEXT,
                ground_or_beans TEXT,
                description_of_taste TEXT,
                price REAL,
                packaging_volume_gram INT)
                """)

    # Читаем файл и лист книги excel
    file_to_read = openpyxl.load_workbook('cofe.xlsx', data_only=True)
    sheet = file_to_read['coffee']

    # Цикл по строкам начиная со второй (в первой заголовки)
    for i in range(2, sheet.max_row + 1):
        # Объявление списка
        data = []
        # Цикл по столбцам от 1 до 7
        for col in range(1, 7 + 1):
            # value содержит значение ячейки с координатами i col
            value = sheet.cell(i, col).value
            # все текстовые данные с заглавной буквы без лишних пробелов
            if 2 <= col <= 5:
                value = value.strip().capitalize()
                value = value.replace('ё', 'е')
            # Список который мы потом будем добавлять
            data.append(value)

        print(data)

        # Вставка данных в поля таблицы
        cursor.execute(f"INSERT INTO Coffee VALUES (?, ?, ?, ?, ?, ?, ?);", (data[0], data[1], data[2], data[3],
                                                                             data[4], data[5], data[6]))

    connect.commit()
    connect.close()


export_to_sqlite()
